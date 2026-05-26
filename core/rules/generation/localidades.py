import json
from pathlib import Path

import pyogrio
from openpyxl import load_workbook

from core.rules.domain_hygiene import build_accepted_values_and_aliases
from core.rules.engine import classify_field_value, load_rule_profile


LOCALIDADES_DOMAIN_COLUMNS = (
    "CD_UF",
    "NM_UF",
    "SIGLA_UF",
    "CD_MUN",
    "NM_MUN",
    "CD_RGINT",
    "NM_RGINT",
    "CD_RGI",
    "NM_RGI",
    "CT_LOCALIDADE",
    "SCT_LOCALIDADE",
)

LOCALIDADES_RELATION_PAIRS = (
    ("cd_uf_to_nm_uf", "sdb_cd_uf", "sdb_nm_uf"),
    ("cd_uf_to_sigla_uf", "sdb_cd_uf", "sdb_sigla_uf"),
    ("cd_mun_to_nm_mun", "sdb_cd_mun", "sdb_nm_mun"),
    ("cd_mun_to_cd_uf", "sdb_cd_mun", "sdb_cd_uf"),
    ("cd_rgint_to_nm_rgint", "sdb_cd_rgint", "sdb_nm_rgint"),
    ("cd_rgint_to_cd_uf", "sdb_cd_rgint", "sdb_cd_uf"),
    ("cd_rgi_to_nm_rgi", "sdb_cd_rgi", "sdb_nm_rgi"),
    ("cd_rgi_to_cd_rgint", "sdb_cd_rgi", "sdb_cd_rgint"),
    ("cd_rgi_to_cd_uf", "sdb_cd_rgi", "sdb_cd_uf"),
)


def generate_localidades_domains(workbook_path, output_path):
    workbook_path = Path(workbook_path)
    values_by_column = read_values_sheet(workbook_path)
    fields = {}

    for column in LOCALIDADES_DOMAIN_COLUMNS:
        accepted_values, aliases = build_accepted_values_and_aliases(
            values_by_column[column]
        )
        fields[sdb(column)] = {
            "accepted_values": accepted_values,
            "aliases": aliases,
        }

    write_json(output_path, {"fields": fields})
    return {
        field: {
            "accepted_values": len(field_rules["accepted_values"]),
            "aliases": len(field_rules["aliases"]),
        }
        for field, field_rules in fields.items()
    }


def generate_localidades_relations(dataset_path, output_path, profile_name):
    profile = load_rule_profile(profile_name)
    original_columns = sorted(
        {
            column[4:].upper()
            for _name, source_column, target_column in LOCALIDADES_RELATION_PAIRS
            for column in (source_column, target_column)
        }
    )
    dataframe = pyogrio.read_dataframe(
        dataset_path,
        columns=original_columns,
        read_geometry=False,
        use_arrow=True,
    )
    dataframe.columns = [sdb(column) for column in dataframe.columns]

    relations = {}
    ambiguous = {}
    cache = {}

    for relation_name, source_column, target_column in LOCALIDADES_RELATION_PAIRS:
        mapping_sets = {}
        deduped = dataframe[[source_column, target_column]].dropna().drop_duplicates()
        for source_value, target_value in deduped.itertuples(index=False, name=None):
            source_clean = clean_value(source_value)
            target_clean = clean_value(target_value)
            if not source_clean or not target_clean:
                continue
            source_norm = normalize_value(profile, source_column, source_clean, cache)
            target_norm = normalize_value(profile, target_column, target_clean, cache)
            if not source_norm or not target_norm:
                continue
            mapping_sets.setdefault(str(source_norm), set()).add(str(target_norm))

        ambiguous_values = {
            key: sorted(values)
            for key, values in mapping_sets.items()
            if len(values) != 1
        }
        if ambiguous_values:
            ambiguous[relation_name] = ambiguous_values

        relations[relation_name] = {
            key: next(iter(values))
            for key, values in sorted(mapping_sets.items(), key=lambda item: item[0])
            if len(values) == 1
        }

    write_json(output_path, {"relations": relations})
    return {
        "relations": {name: len(mapping) for name, mapping in relations.items()},
        "ambiguous": {name: len(mapping) for name, mapping in ambiguous.items()},
    }


def read_values_sheet(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["valores"]
    headers = [str(cell).strip() for cell in next(worksheet.iter_rows(values_only=True))]
    values_by_column = {header: [] for header in headers}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for header, value in zip(headers, row):
            values_by_column[header].append(value)

    return values_by_column


def normalize_value(profile, column, value, cache):
    key = (column, value)
    if key not in cache:
        cache[key] = classify_field_value(profile, column, value)["normalized_value"]
    return cache[key]


def clean_value(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "<na>"}:
        return None
    return text


def sdb(column):
    column_text = str(column).strip().lower()
    if column_text.startswith("sdb_"):
        return column_text
    return f"sdb_{column_text}"


def write_json(path, data):
    Path(path).write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


__all__ = [
    "LOCALIDADES_DOMAIN_COLUMNS",
    "LOCALIDADES_RELATION_PAIRS",
    "generate_localidades_domains",
    "generate_localidades_relations",
]
